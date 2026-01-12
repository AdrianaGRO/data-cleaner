import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment



def remove_duplicates(df, subset_columns=None, keep_rule="first"):
    """
    Remove duplicate rows from the DataFrame on selected columns.

    Parameters:
    df (pd.DataFrame): Input DataFrame.
    subset_columns (list or None): Columns to consider for identifying duplicates.
    keep_rule (str or False): 'first', 'last', or False.

    Returns:
    tuple: (cleaned_df, removed_count)
    """
    allowed_keep_rules = ["first", "last", False]
    if keep_rule not in allowed_keep_rules:
        raise ValueError(f"Invalid keep_rule '{keep_rule}'. Allowed values: {allowed_keep_rules}")

    if subset_columns is not None:
        for col in subset_columns:
            if col not in df.columns:
                raise ValueError(f"Column '{col}' not found. Available columns: {list(df.columns)}")

    initial_count = len(df)
    cleaned_df = df.drop_duplicates(subset=subset_columns, keep=keep_rule)
    removed_count = initial_count - len(cleaned_df)

    return cleaned_df, removed_count


def merge_excel_files(file_paths, sheet_name=0):
    """ Merge multiple Excel or CSV files (or process a single file).
    
    Supports:
    - .xlsx (Excel 2007+)
    - .xls (Excel 97-2003)
    - .csv (Comma-separated values)
    
    Parameters:
    file_paths (list): List of file paths to merge (can be 1 or more)
    sheet_name (int or str): Sheet name or index to read (Excel only)
    
    Returns:
    pd.DataFrame: Merged DataFrame"""
    
    if not file_paths:
        raise ValueError("file_paths cannot be empty")
    
    # Handle single file (no merging needed)
    if len(file_paths) == 1:
        file_path = file_paths[0]
        
        #Check file extension
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        return df
    
    
    # Handle multiple files (original merging logic)
    dfs = []
    first_columns = None
    
    for i, path in enumerate(file_paths):
        #Read file based on extension
        if path.lower().endswith('.csv'):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path, sheet_name=sheet_name)
        
        #Check column consistency
        
        if i == 0:
            first_columns = list(df.columns)
        else:
            if list(df.columns) != first_columns:
                raise ValueError(
                    f"Column mismatch: {path} has different columns than first file.\n"
                    f"Expected: {first_columns}\n"
                    f"Got: {list(df.columns)}"
                )
        
        dfs.append(df)
    #Merge all types of files (Excel and CSV)
    merged_df = pd.concat(dfs, ignore_index=True)
    return merged_df


def clean_excel_pipeline(file_paths, subset_columns=None, keep_rule="first", sheet_name=0):
    """
    Complete pipeline to merge Excel files (deduplication handled separately).

    Parameters:
    file_paths (list): List of Excel files to merge.
    subset_columns (list): Columns to consider for duplicate removal (not used, kept for compatibility).
    keep_rule (str/bool): Duplicate removal rule (not used, kept for compatibility).
    sheet_name (str/int): Sheet name or index.

    Returns:
    tuple: (final_df, summary) where summary contains operation metadata.
    """
    merged_df = merge_excel_files(file_paths, sheet_name=sheet_name)
    
    rows_before = len(merged_df)
    
    # No deduplication here - will be done separately after blank row removal
    final_df = merged_df
    
    rows_after = len(final_df)
    
    summary = {
        "files_merged": len(file_paths),
        "rows_before": rows_before,
        "rows_after": rows_after,
        "duplicates_removed": 0  # Will be calculated separately
    }

    return final_df, summary

def standardize_dates(df, date_columns):
    """
    Standardize date columns to YYYY-MM-DD format.
    
    Parameters:
    df (pd.DataFrame): Input DataFrame
    date_columns (list): Column names containing dates
    
    Returns:
    tuple: (cleaned_df, conversion_count)
    """
    if not date_columns:
        raise ValueError("date_columns cannot be empty")
    
    df_copy = df.copy()
    conversion_count = 0
    
    for col in date_columns:
        if col not in df_copy.columns:
            raise ValueError(f"Column '{col}' not found. Available columns: {list(df_copy.columns)}")
        
        def parse_date(date_str):
            """Try multiple date formats."""
            if pd.isna(date_str):
                return pd.NaT
            
            # List of common date formats to try
            formats = [
                '%m/%d/%Y',      # 01/15/2024
                '%Y-%m-%d',      # 2024-02-20
                '%B %d, %Y',     # March 10, 2024
                '%b %d, %Y',     # Mar 10, 2024
                '%d/%m/%Y',      # 15/01/2024
                '%Y/%m/%d',      # 2024/01/15
            ]
            
            for fmt in formats:
                try:
                    return pd.to_datetime(date_str, format=fmt)
                except (ValueError, TypeError):
                    continue
            
            # If all formats fail, try pandas auto-detection
            try:
                return pd.to_datetime(date_str)
            except:
                return pd.NaT
        
        # Apply date parsing
        df_copy[col] = df_copy[col].apply(parse_date)
        
        # Count valid dates
        valid_dates = df_copy[col].notna().sum()
        conversion_count += valid_dates
        
        # Format as YYYY-MM-DD string
        df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d')
        
        # Replace NaN with empty string
        df_copy[col] = df_copy[col].fillna('')
    
    return df_copy, conversion_count

def clean_phone_numbers(df, phone_columns):
    """
    Clean and standardize phone numbers to (XXX) XXX-XXXX format.
    Handles extensions, country codes, and various delimiters.
    
    Parameters:
    df (pd.DataFrame): Input DataFrame
    phone_columns (list): Column names containing phone numbers
    
    Returns:
    tuple: (cleaned_df, cleaned_count)
    
    """
    
    if not phone_columns:
        raise ValueError("phone_columns cannot be empty")
    
    df_copy = df.copy()
    cleaned_count = 0 
    
    for col in phone_columns:
        if col not in df_copy.columns:
            raise ValueError(f"Column '{col}' not found. Available columns: {list(df_copy.columns)}")
        
        def format_phone(phone):
            if pd.isna(phone):
                return ""
            
            phone_str = str(phone)
            
            #Remove extension
            if 'x' in phone_str.lower():
                phone_str = phone_str.lower().split('x')[0]
            
            #Remove all non-digit characters
            digits = re.sub(r'\D', '', phone_str)                       
            
            #Handle different lengths
            if len(digits) == 10:
                # Standard US format
                return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
            elif len(digits) == 11 and digits[0] == '1':
                # Remove leading 1 (US country code)
                return f"({digits[1:4]}) {digits[4:7]}-{digits[7:11]}"
            elif len(digits) > 11:
                # Take last 10 digits (handles long extensions)
                digits = digits[-10:]
                return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
            else:   
                #Invalid length, return empty
                return ""
            
        #Apply formatting to column
        df_copy[col] = df_copy[col].apply(format_phone)
        
        #Count successfully formatted numbers
        cleaned_count += df_copy[col].str.match(r'^\(\d{3}\) \d{3}-\d{4}$', na=False).sum()
        
    return df_copy, cleaned_count
        
            
            
def remove_blank_rows(df):
    """
    Remove rows where all cells are empty or contain only whitespace.
    
    Parameters:
    df (pd.DataFrame): Input DataFrame
    
    Returns:
    tuple: (cleaned_df, removed_count)
    
    """
    
    initial_count = len(df)
    
    df_copy = df.copy()
    
    # Enhanced blank detection
    # 1. Replace empty strings with NaN
    df_copy = df_copy.replace("", pd.NA)
    
    # 2. Replace whitespace-only strings with NaN
    df_copy = df_copy.replace(r'^\s*$', pd.NA, regex=True)
    
    # 3. Replace common "empty" values including 'nan' strings
    df_copy = df_copy.replace(['None', 'null', 'NULL', 'NaN', 'nan', 'Nan', 'NAN'], pd.NA)
    
    # 4. Handle cases where all columns contain 'nan' as string
    nan_mask = (df_copy.astype(str) == 'nan').all(axis=1)
    df_copy.loc[nan_mask] = pd.NA
    
    # 4. Drop rows where all elements are NaN
    df_copy = df_copy.dropna(how='all')
    
    # Reset index
    df_copy = df_copy.reset_index(drop=True)
    
    removed_count = initial_count - len(df_copy)
    
    return df_copy, removed_count

def format_excel_output(file_path):
    """Format excel file with bold headers and auto column width..
    
    Parameters:
    file_path (str): Path to the Excel file.
    
    returns:
    None(modifies file in place)
    
    """
    
    #Load workbook and select active sheet
    wb = load_workbook(file_path)
    ws = wb.active
    
    #Make header row bold
    
    for cell in ws[1]: 
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    #Auto_adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
            
            
        #Set the width(add padding)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
        
    #Save changes
    wb.save(file_path)
    